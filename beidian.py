import time
import os
import datetime
import json
from flask import Flask,send_file,render_template
from flask_apscheduler import APScheduler
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from random import randint

class Config(object):
    JOBS = [
            {
               'id':'job1',
               'func':'beidian:get_data',
               'args': '',
               'trigger': {
                    'type': 'cron',
                    'hour':'0-23',
                    'minute':'32'
                }

             }
        ]

    SCHEDULER_API_ENABLED = True


headers = {'User-Agent':"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1"}
urls = ['https://m.beidian.com/detail/detail.html?iid=31022080&shop_id=1480734&utm_source=bd_spfzlj&offer_code=0&r_uid=NzIzMzUxOTc&r_uid=NzIzMzUxOTc','https://m.beidian.com/detail/detail.html?iid=30409280&shop_id=1480734&utm_source=&offer_code=0','https://m.beidian.com/detail/detail.html?iid=32022908&shop_id=1480734&utm_source=bd_spfxhy&offer_code=0&r_uid=NzIzMzUxOTc&r_uid=NzIzMzUxOTc']


service_args=[]
service_args.append('--load-images=no')
service_args.append('--disk-cache=yes')
service_args.append('--ignore-ssl-errors=true')




col1 = ["C","I","O"]
col2 = ["D","J","P"]
col3 = ["E","K","Q"]

col4 = ["F","L","R"]

app = Flask(__name__)
scheduler = APScheduler()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download')
def download():
    return send_file('static/sales_data.xlsx',as_attachment=True)

@app.route('/pause')
def pause():
    scheduler.pause_job('job1')
    return render_template('resume.html')


@app.route('/resume')
def resume():
    scheduler.resume_job('job1')
    return render_template('pause.html')


def get_data():
    driver=webdriver.PhantomJS(executable_path='D:/App/phantomjs/bin/phantomjs',service_args=service_args)
    sales = []
    stocks = []
    for url in urls:
        driver.get(url)
        time.sleep(randint(1,3))
        try:
            element = WebDriverWait(driver,30).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'J_sellerCount')))
        finally:
            sales_number = driver.find_element_by_class_name('J_sellerCount').text
            stocks_number = driver.find_element_by_class_name('J_stockNum').text
        sales_number = sales_number.replace('人已买','')
        stocks_number = stocks_number.replace('库存剩','').replace('件','')
        sales.append(sales_number)
        stocks.append(int(stocks_number))
    driver.quit()
    print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sales,stocks)

    if not os.path.exists('static'):
        os.makedirs('static')

    filename_path = 'static/sales_data.xlsx'

    f = open('var.json','r')
    var = json.loads(f.read())
    f.close()



    wb = load_workbook(filename=filename_path)
    ws = wb.active



    for i in range(len(sales)):
        ws.cell(column = 1,row = var['row1'],value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(column = 2+6*i,row = var['row1'],value = sales[i])

    for i in range(len(stocks)):
        ws.cell(column = 3+6*i,row = var['row2'],value = stocks[i])



    if var['row1'] >= 3:
        for i in range(len(sales)):
            f = "={}{}-{}{}".format(col1[i],str(var['row3']-1),col1[i],str(var['row3']))
            ws.cell(column = 4+6*i,row = var['row3'],value = f)
        var['row3'] += 1


    if var['row1'] >= 4:
        for i in range(len(sales)):
            f = "={}{}-{}{}".format(col2[i],str(var['row4']),col2[i],str(var['row4']-1))
            ws.cell(column = 5+6*i,row = var['row4'],value = f)
        var['row4'] += 1

    if var['row1'] >= 5:
        for i in range(len(sales)):
            a = "{}{}".format(col1[i],var['row1']-2)
            b = "{}{}".format(col1[i],var['row1']-1)
            c = "{}{}".format(col1[i],var['row1']-3)
            d = "{}{}".format(col1[i],var['row1']-2)

            e = "{}{}".format(col4[i],var['row1'])
            align = Alignment(horizontal='right',vertical='center',wrap_text=True)
            if (ws[a].value-ws[b].value)-(ws[c].value-ws[d].value) == 0:
                ws.cell(column = 6+6*i,row = var['row5'],value = "--")
                ws[e].alignment = align

            else:
                f = "={}{}-{}{}/{}{}".format(col3[i],str(var['row5']),col3[i],str(var['row5']-1),col3[i],str(var['row5']-1))
                ws.cell(column = 6+6*i,row = var['row5'],value = f)
                ws[e].number_format = '0.00%'
        var['row5'] +=1


    var['row1'] += 1
    var['row2'] += 1


    wb.save(filename=filename_path)

    f = open('var.json','r+')
    f.write(json.dumps(var))
    f.close()

def my_listener(event):
    if event.exception:
        print('The job crashed :(')
    else:
        print('The job worked :)')

