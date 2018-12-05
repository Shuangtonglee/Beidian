from apscheduler.schedulers.blocking import BlockingScheduler
import time
import os
import datetime
from openpyxl import load_workbook,Workbook
from selenium import webdriver
from random import randint

service_args=[]
service_args.append('--load-images=no')
service_args.append('--disk-cache=yes')
service_args.append('--ignore-ssl-errors=true')
sched = BlockingScheduler()

urls = ['https://ibei.cn/1y5uOZ','https://m.beidian.com/detail/detail.html?iid=30409280&shop_id=1480734&utm_source=&offer_code=0']

@sched.scheduled_job('interval', seconds=60)
#@sched.scheduled_job('cron', hour='0-23',minute='00')
def get_data():
    # driver=webdriver.PhantomJS(executable_path='bin/phantomjs',service_args=service_args)
    # sales_number_list = []
    # for url in urls:
    #     driver.get(url)
    #     time.sleep(randint(1,3))
    #     sales_number = driver.find_element_by_class_name('J_sellerCount').text
    #     sales_number_list.append(sales_number)
    # driver.quit()
    # print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sales_number_list)

    if not os.path.exists('static'):
        os.makedirs('static')

    filename_path = 'static/sales_data.xlsx'
    sheet_name = datetime.datetime.now().strftime('%m-%d')
    if not os.path.exists(filename_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(['时间','销量'])
    else:
        wb = load_workbook(filename=filename_path)
        sheets = wb.get_sheet_names()
        if not sheet_name in sheets:
            ws = wb.create_sheet(sheet_name)
            ws.append(['时间','销量'])
        else:
            ws = wb[sheet_name]


    for sale in [1,3]:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),sale)
        ws.append([datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),sale])
    wb.save(filename=filename_path)

sched.start()